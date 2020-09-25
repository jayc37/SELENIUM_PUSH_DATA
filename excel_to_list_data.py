#! coding: utf-8

import re
import string
import os,socket
import sys
import codecs
import pyodbc
import logging
import argparse
from threading import Timer
from time import sleep
from random import randint
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

def sleeping(min_s, max_s):
    sleep(randint(min_s, max_s))

class pushdata:
    def __init__(self,servername,dbname,uid,pwd):
        self.username = ''
        self.password = ''
        self.driver = webdriver.Chrome(ChromeDriverManager().install(),options=op)
        self.conn = pyodbc.connect('DRIVER={SQL Server};SERVER=%s;DATABASE=%s;UID=%s;PWD=%s' % (servername, dbname, uid, pwd)) 
        self.curr = self.conn.cursor()
        self.sokhung = ''
        self.file_name = ''
        date_n = datetime.now()
        date_n = date_n.strftime("%d_%m")
        self.date = date_n
        self.name_ch = ''
        self.err_name = ''
        self.point = 0
        self.icount = 0
        self.data_from_excel_file = None
        self.err_list = []
        self.not_match = []

    def force_timeout(self):
        try:
            print('kill chrome')
            os.system('taskkill /im chrome.exe /f /t') #kill chrome
            self.driver = webdriver.Chrome(ChromeDriverManager().install(),options=op) #create new chrome
            self.login()
            self.get_object()

        except:
            print('kill error')
            err = 'STOP at line ' + str(self.point) + ' trong file excel'
            print(err)
            self.write_log_thatbai(err)

    def login(self):
            t = Timer(60, self.force_timeout)
            t.start()
            """ LOGIN """
            UURL = ""
            self.driver.get(UURL)
            self.driver.maximize_window()
        # (k) : khối dữ liệu lấy từ file excel
        #----<Khối lệnh này điền dữ liệu vào form đăng nhập trên trang HMS>-------
                #Tìm kiếm label và điền vào username-login
            try:
                actions = ActionChains(self.driver)
                actions.send_keys(Keys.F5).perform() # Refresh lại trang trước khi đăng nhập
                username_ele = self.driver.find_element_by_name("SWEUserName") 
                username_ele.send_keys(self.username) # điền vào username
                sleeping(1,2)
                
                #Tìm kiếm label và điền vào password-login
                password_ele = self.driver.find_element_by_name("SWEPassword")
                password_ele.send_keys(self.password) # điền pwd vào label pwd
                sleeping(1,2)
                
                #Tìm kiếm và ấn nút login
                login_ele = self.driver.find_element_by_xpath(".//*[@id='s_swepi_22']")
                login_ele.click()
                sleeping(3, 5) #--

            except:
                print('login fail')
                pass   
            t.cancel()
    def write_log_error(self,message):
        """
            File log error
        """
        with open('log\\logfile_error', 'a',encoding="utf-8") as f:
            da = datetime.now()
            da = da.strftime("%d/%m/%Y %H:%M:%S")
            f.write(str(message) + '----> at: ' + str(da))
            f.seek(0,0)
        f.close()

    def getacount(self,data_from_excel,name,file_name):
        try:
            self.file_name = file_name
            self.name_ch = name
            self.data_from_excel_file = data_from_excel
            self.curr.execute("SELECT username,passwords from user_HMS_write where cuahang = '{}'".format(str(name)))
            result = self.curr.fetchall()
            for i in range(len(result)):
                self.username = result[i][0]
                self.password = result[i][1]
                self.login()
                self.get_object() #Gửi (k) truyền vào hàm xử lý
            print('done')
        except:
            print(str(name))

    def kiemtra(self,sokhung,somay,sodienthoai,biensoxe):

        sk = re.search(r'^(.{17})$',sokhung)
        try:
            if sk:
                self.send_values_sokhung(sokhung,somay,sodienthoai,biensoxe)
            elif somay != 'None':
                self.send_values_somay(somay,sodienthoai,biensoxe)
            elif biensoxe != 'None':
                self.send_values_biensoxe(sodienthoai,biensoxe)
            
        except socket.error:
            self.write_error_number_excels()
            pass   
    #----<Khối lệnh này xử lý và kiểm tra từng dòng dữ liệu từ (k)>-------
    def get_object(self):
        """ PROCESS """
        try:
            sleeping(1,2)
             # Mảng này chứa các dòng dữ liệu bị sai của (k)
            dat_ex = self.data_from_excel_file
            for i in range(len(dat_ex)):
                self.icount = i
                if self.point != self.icount:
                    eror_x = dat_ex[int(i) - 1]
                    self.err_name = 'dòng dữ liệu bị bỏ qua'
                    eror_x.append(self.err_name)
                    self.not_match.append(dat_ex[int(i) - 1]) 
                    self.err_list.append(eror_x)
                dx =  dat_ex[i] # mảng này chứa dòng dữ liệu bao gồm: số khung,số máy,số điện thoại, biển số xe
                sokhung = str(dx[0])
                self.sokhung = sokhung
                somay = str(dx[1]) 
                sodienthoai = str(dx[2])
                biensoxe =  str(dx[3])
                checknumphone = re.search(r'(([0-9]{10})\b)|(([0-9]{11})\b)',str(sodienthoai)) # Biến này kiểm tra xem độ dài của số điện thoại có đủ 10 hoặc 11 kí tự hay không
                valip = re.search(r'^0',sodienthoai) #Biến này kiểm tra xem số điện thoại có bắt đầu bằng số 0
                chieudaisdt = len([i for i in range(len(sodienthoai))])
                t = Timer(120, self.force_timeout)
                t.start()

                if checknumphone:
                    self.kiemtra(sokhung,somay,sodienthoai,biensoxe)#truyền đi dòng dữ liệu đúng 
                elif valip: 
                    print('kiem tra lai so dien thoai:  "'
                    + str(sodienthoai) + '" at:\nSokhung: '+ str(sokhung) + '\nSomay: ' + str(somay) + '\nBiensoxe: ' + str(biensoxe))
                    eror_x = dx
                    self.err_name = 'dòng dữ liệu sai'
                    eror_x = dx.append(self.err_name)
                    self.err_list.append(eror_x) # ghi lại dòng dữ liệu sai
                    self.point += 1
                elif chieudaisdt == 9:
                    sodienthoai = '0' + str(sodienthoai) #thêm số 0 vào đầu số nếu số có: 9 kí tự và k bắt đầu bằng 0
                    self.kiemtra(sokhung,somay,sodienthoai,biensoxe)

                else:
                    eror_x = dx
                    self.err_name = 'dòng dữ liệu sai'
                    eror_x = dx.append(self.err_name)
                    self.err_list.append(eror_x) # ghi lại dòng dữ liệu sai
                    self.point += 1
                t.cancel()
                
        except Exception as e:
            logging.error('<get_object function> ==>::' + str(e),exc_info=True)
            message = '<get_object function> ==>:: + {}'.format(str(e))
            eror_x = self.data_from_excel_file[self.icount]
            self.err_name = 'lỗi trang chưa được load'
            eror_x.append(self.err_name)
            self.err_list.append(eror_x) # ghi lại dòng dữ liệu sai
            self.write_log_error(message)
            pass
        if len(self.err_list) != 0:
            self.write_error_number_excels()
        if len(self.not_match) != 0:
            for i in range(len(self.not_match)):
                log = self.not_match[i]
                message = 'Ghi khong thanh cong: ' + str(log[0]) + ' - ' + str(log[1]) + ' - ' + str(log[2])
                self.write_log_thatbai(message)
    #khối lệnh này ghi mảng err_list vào file excel chứa trong folder error_file
    def write_error_number_excels(self):
        try:
            pl = self.err_list
            ws = Workbook()
            wb = ws.active
            wb['A1'] = 'SOKHUNG'
            wb['B1'] = 'SOMAY'
            wb['C1'] = 'SODIENTHOAISAI'
            wb['D1'] = 'BIENSOXE'
            wb['E1'] = 'CUAHANG'
            wb['F1'] = 'ERROR_LOG'
            for i in range(len(pl)):
                if pl[i] is not None:
                    w = 2
                    w +=i
                    wb['A{}'.format(w)] =  pl[i][0]
                    wb['B{}'.format(w)] =  pl[i][1]
                    wb['C{}'.format(w)] =  pl[i][2]
                    wb['D{}'.format(w)] =  pl[i][3]
                    wb['E{}'.format(w)] =  self.name_ch
                    wb['F{}'.format(w)] =  pl[i][4]
                    self.write_log_thatbai(pl[i])   
            ws.save('error_file\\errorr.xlsx')
            print("CO MOT SO DONG BI SAI DUOC GHI VAO error_file/error.xlsx")
        except Exception as e:
            print(str(e))
    #Khối lệnh này tìm dòng dữ liệu trên trang bằng số khung
    def send_values_sokhung(self,sokhung,somay,sodienthoai,biensoxe):
        """ PROCESS SOKHUNG"""
        try:
            thongtinxe = self.driver.find_element(By.XPATH,'//a[text()="Thông tin xe"]')
            thongtinxe.click()
            sleeping(1, 2)
                        
            dealerall_button = self.driver.find_element_by_name("s_vis_div")
            dealerall_button.click()

            dealerall_option = self.driver.find_element(By.XPATH,"//option[text()='All Visible Vehicle']")
            dealerall_option.click()
            sleeping(1, 2) 
            #type in sokhung
            
            sokhung_label = self.driver.find_element_by_name("s_1_1_102_0")
            sokhung_label.send_keys(sokhung)
            # sleeping(1,2)
            sokhung_label.send_keys(Keys.ENTER)
            sleeping(1, 2) 
            # click
            find_sdt_column = self.driver.find_element_by_id("jqgh_s_1_l_Cellular_Phone__")
            find_sdt_column.click()
            find_sdt_label = self.driver.find_element_by_id("1_s_1_l_Cellular_Phone__")
            actions = ActionChains(self.driver)
            actions.double_click(find_sdt_label).perform()
            sleeping(1,2)
            actions.send_keys(sodienthoai).perform()
            actions.send_keys(Keys.ENTER).perform()
            log_line = 'su dung sokhung: ' + str(sokhung) +' luu ' +str(sodienthoai)+ ' thanh cong '              
            date = datetime.now()
            date = date.strftime("%d/%m/%Y %H:%M:%S")
            self.write_log_thanhcong(date,log_line)
            self.point += 1
            # clickfindfield = self.driver.find_element_by_xpath("//*[@id='s_S_A1_div']/form/span/div/div[1]")
            # clickfindfield.click()
            # sdt = self.driver.find_element_by_xpath("//*[@id='1_s_1_l_Cellular_Phone__']")

            # date = datetime.now()
            # date = date.strftime("%d/%m/%Y %H:%M:%S")
            # self.date = date
            # by = 'sokhung'
            # return by
            sleeping(2,3)
            
        except Exception as e:
            print('sokhung sai')
            print('...........')
            print('tim bang somay')
            self.send_values_somay(somay,sodienthoai,biensoxe)
            pass
    #Khối lệnh này tìm dòng dữ liệu trên trang hms bằng số máy
    def send_values_somay(self,somay,sodienthoai,biensoxe):
        """ PROCESS SOMAY"""
        try:
            thongtinxe = self.driver.find_element(By.XPATH,'//a[text()="Thông tin xe"]')
            thongtinxe.click()
            sleeping(3, 5)
                        
            dealerall_button = self.driver.find_element_by_name("s_vis_div")
            dealerall_button.click()

            dealerall_option = self.driver.find_element(By.XPATH,"//option[text()='All Visible Vehicle']")
            dealerall_option.click()
            sleeping(3, 5) 
            #type in somay
            
            so_may = self.driver.find_element_by_name("s_1_1_93_0")
            so_may.send_keys(somay)
            # sleeping(1,2)
                        
            so_may.send_keys(Keys.ENTER)
            sleeping(3, 5) 
            # click
            find_sdt_column = self.driver.find_element_by_id("jqgh_s_1_l_Cellular_Phone__")
            find_sdt_column.click()
            find_sdt_label = self.driver.find_element_by_id("1_s_1_l_Cellular_Phone__")
            find_sdt_label.click()
            actions = ActionChains(self.driver)
            actions.double_click(find_sdt_label).perform()
            sleeping(1,2)
            actions.send_keys(sodienthoai).perform()
            actions.send_keys(Keys.ENTER).perform()
            log_line = 'su dung somay: ' + str(somay) +' luu ' +str(sodienthoai)+ ' thanh cong '              
            date = datetime.now()
            date = date.strftime("%d/%m/%Y %H:%M:%S")

            self.write_log_thanhcong(date,log_line)
            self.point += 1

            sleeping(2,3)
            
        except Exception as e:
            print('somay sai')
            print('...........')
            print('tim bang biensoxe')
            self.send_values_biensoxe(sodienthoai,biensoxe)
                # by = 'biensoxe'
                # return by
    #Khối lệnh này tìm dòng dữ liệu trên trang hms bằng biển số xe
    def write_log_thanhcong(self,ngayghilog,log_line):
        print(str(log_line))
        with open('log\\log_thanhcong','a') as f:
            f.write(str(ngayghilog)+ ' :  '+ str(log_line) +'\n')
            f.seek(0,0)
        f.close()

    def write_log_thatbai(self,log_line):
        print(str(log_line))
        with open('log\\log_thatbai','a') as f:
            f.write(str(log_line) +'\n')
            f.seek(0,0)
        f.close()

    def send_values_biensoxe(self,sodienthoai,biensoxe):
        """ PROCESS """
        try:

            thongtinxe = self.driver.find_element_by_xpath("//*[@id='s_sctrl_tabScreen']/ul/li[7]")
            thongtinxe.click()
            sleeping(3, 5)
                        
            dealerall = self.driver.find_element_by_name("s_vis_div")
            dealerall.click()

            dealerall2 = self.driver.find_element_by_xpath("//*[@id='s_vis_div']/select/option[1]")
            dealerall2.click()
            sleeping(3, 5) 

            bienso = self.driver.find_element_by_name("s_1_1_103_0")
            bienso.send_keys(biensoxe)
            # sleeping(1,2)
                        
            bienso.send_keys(Keys.ENTER)
            sleeping(3, 5) 
             # click
            actions = ActionChains(self.driver)
            find_sdt_column = self.driver.find_element_by_id("jqgh_s_1_l_Cellular_Phone__")
            find_sdt_column.click()
            find_sdt_label = self.driver.find_element_by_id("1_s_1_l_Cellular_Phone__")
            find_sdt_label.click()
            actions.double_click(find_sdt_label).perform()
            sleeping(1,2)
            actions.send_keys(sodienthoai).perform()
            
            actions.send_keys(Keys.ENTER).perform()
            log_line = 'su dung biensoxe: ' + str(somay) +' luu ' +str(sodienthoai)+ ' thanh cong '              
            date = datetime.now()
            date = date.strftime("%d/%m/%Y %H:%M:%S")

            self.write_log_thanhcong(date,log_line)
            self.point += 1

            sleeping(2,3)
            
        except Exception as e:
            self.err_name = 'sai sokhung,somay,biensoxe'
            x = self.data_from_excel_file[self.icount]
            x.append(self.err_name)
            self.err_list.append(x)
            self.point += 1
            pass
#khối lệnh này chuyển dữ liệu từ #file_excel chuyển thành mảng
def work_sheet(wsheet):
    data_sheet = []
    col = [] #column in sheet
    for c in range(wsheet.max_column):
        #got alphabels with max_(len)_column found in worksheet
        col.append(string.ascii_uppercase[c])

    for r in range(2,wsheet.max_row + 1):
        data_row = []
        for c in range(len(col)):
            #got values exactly with "sheet[colum-row]"
            data = wsheet['{}{}'.format(col[c],r)].value
            data_row.append(data)
        data_sheet.append(data_row)
    return data_sheet

def get_data(wbook):
    data = []
    # got values with multiple sheet in workbook
    for sheet in wbook.worksheets:
        wsheet = wbook.active
        data_book = work_sheet(sheet)
        data.extend(data_book)
    return data

if __name__ == '__main__':
    with open('src\\SQL_server_connection.txt','r') as f:
        x = f.readlines()
        servername = str(x[0]).replace('\n','')
        dbname = str(x[1]).replace('\n','')

        uid = str(x[2]).replace('\n','')
        pwd = str(x[3]).replace('\n','')
    f.close()
    with open('log\\log_thanhcong','w') as f:
        f.truncate()
    f.close()
    with open('log\\logfile_error','w') as f:
        f.truncate()
    f.close()
    with open('log\\log_thatbai','w') as f:
        f.truncate()
    f.close()

    op = webdriver.ChromeOptions()
    op.add_argument("--disable-gpu")
    op.add_argument('start-maximized')
    op.add_argument("--disable-extensions")
    op.add_argument("enable-automation")
    op.add_argument("--no-sandbox")
    op.add_argument("--disable-infobars")
    op.add_argument("--disable-dev-shm-usage")
    op.add_argument("--disable-browser-side-navigation")
    op.add_experimental_option("excludeSwitches", ["enable-logging"])

    pushda = pushdata(servername,dbname,uid,pwd)
    # Tìm kiếm các file excel trong folder file_excel_here
    for root, dirs, files in os.walk("file_excel_here"):
        for file in files: 
                if file.endswith(".xlsx") and '~' not in file: #tìm ra file excel
                    x = re.search(r'pt\d.*',file) #so khớp
                    file_name = os.path.join(root, file)
                    wb = load_workbook(file_name)
                    if x:
                        data = get_data(wb)
                        name = re.sub('.xlsx','',x.group()) # cắt chuỗi có kí tự 'pt'+number
                        pushda.getacount(data,name,file) # đăng nhập
                        wb.close()
                    else:
                        print("name_of_file must end with 'pt' + 'number' like 'example_pt1'")
    pushda.driver.close()
